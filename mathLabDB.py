import cv2
import string
import openpyxl

def loadDataBank():

    path = 'DB/'

    wb = openpyxl.Workbook()  # Crear archivo excel

    ABCD = list(string.ascii_uppercase)  # matriz con todas las letras del abcedario

    countRow = 1
    countCol = 1

    for letter in ABCD:
        sheet = wb.active
        for num in range(10):
            countCol = 1
            pathImg = '{}{}{}.png'.format(path, letter, num + 1)
            print(pathImg)
            img = cv2.imread(pathImg, 2)
            ret, bw_img = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)  # sistema de quemado de imagen gray
            bw = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)  # vectorizado
            # ----------------------------------Guardado del banco de datos dentro del Excel------------------------------
            for row in bw_img:
                countData = 0
                for data in row:
                    sheet.cell(row=countRow, column=countCol, value=data)
                    countData += data / 255
                sheet.cell(row=countRow, column=countCol, value=countData)
                countCol += 1
            countRow += 1

    print("termine :v")
    wb.save('files/DB.xlsx')


