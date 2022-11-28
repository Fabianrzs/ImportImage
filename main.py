import cv2
import string
import openpyxl

path = 'DataBank/'

wb = openpyxl.Workbook()  # Crear archivo excel

ABCD = list(string.ascii_uppercase)  # matriz con todas las letras del abcedario

for letter in ABCD:
    print("Voy por el Directorio: ", letter)

    sheet = wb.create_sheet(letter)
    countRow = 1
    countCol = 1

    for item in range(25):

        pathImg = '{}{}/{}.JPG'.format(path, letter, item + 1)

        print(pathImg)

        img = cv2.imread(pathImg, 2)
        ret, bw_img = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)  # sistema de quemado de imagen gray
        bw = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)  # vectorizado

        # ----------------------------------Guardado del banco de datos dentro del Excel------------------------------

        print("Largo ", len(bw_img))
        countCol = 1
        for row in bw_img:
            print("Ancho ",len(row))
            countData = 0
            for data in row:
                sheet.cell(row=countRow, column=countCol, value=data)
                countData += data/255
            sheet.cell(row=countRow, column=countCol, value=countData)
            countCol += 1
        countRow += 1

print("termine :v")
wb.save('probe.xlsx')

