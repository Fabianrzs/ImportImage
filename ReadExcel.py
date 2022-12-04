import string
import openpyxl

pathInit = 'files/'

def readfiles(path, val):
    ABCD = list(string.ascii_uppercase)  # matriz con todas las letras del abcedario
    wb = openpyxl.load_workbook(pathInit+path)

    sheet = wb.active
    max_col = sheet.max_column
    # wb[a]
    print(wb["A"])
    excel_data = []

    for letter in ABCD:
        sheet = wb[letter]
        max_col = sheet.max_column
        print("Voy por la ", letter)
        for i in range(25):
            row = []
            for j in range(1, max_col + 1):
                cont = i+1
                if (val == 0):
                    cont = 1
                cell_obj = sheet.cell(row=cont, column=j)
                row.append(cell_obj.value)
            excel_data.append(row)
    return excel_data


DataBank= readfiles("DataBank.xlsx", 1)

print("DataBank ", DataBank)

validations = readfiles("Validations.xlsx", 0)

print("Validaciones ", validations)
